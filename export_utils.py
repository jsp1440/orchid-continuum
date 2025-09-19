"""
Export utilities for Excel and CSV downloads
"""
import os
import csv
import json
from io import BytesIO, StringIO
from datetime import datetime
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import current_app

from models import OrchidRecord, UserUpload, User, JudgingAnalysis

class OrchidDataExporter:
    """Export orchid data to various formats"""
    
    def __init__(self):
        self.export_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def export_to_excel(self, records: List[OrchidRecord], user_id: Optional[int] = None, 
                       include_images: bool = False, include_judging: bool = False) -> BytesIO:
        """
        Export orchid records to Excel format
        
        Args:
            records: List of OrchidRecord objects
            user_id: User ID for filtering (None for admin export)
            include_images: Whether to include image data
            include_judging: Whether to include judging analysis
            
        Returns:
            BytesIO object with Excel data
        """
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orchid Records"
        
        # Define headers
        headers = [
            'ID', 'User ID', 'Plant ID', 'Display Name', 'Scientific Name',
            'Genus', 'Species', 'Author', 'Region', 'Native Habitat',
            'Bloom Time', 'Growth Habit', 'Climate Preference',
            'Light Requirements', 'Temperature Range', 
            'Image Filename', 'Photographer', 'Source',
            'AI Description', 'AI Confidence', 'Ingestion Source',
            'Validation Status', 'Featured', 'View Count',
            'Created Date', 'Updated Date'
        ]
        
        if include_judging:
            headers.extend([
                'Judging Scores', 'Award Level', 'Award Worthy',
                'Last Judged', 'Judging Organization'
            ])
        
        # Style headers
        self._style_headers(ws, headers)
        
        # Add data rows
        for row_num, record in enumerate(records, start=2):
            # Get associated upload info
            upload = UserUpload.query.filter_by(orchid_id=record.id).first()
            user = User.query.get(record.user_id) if record.user_id else None
            
            row_data = [
                record.id,
                user.user_id if user else '',
                upload.plant_id if upload else '',
                record.display_name or '',
                record.scientific_name or '',
                record.genus or '',
                record.species or '',
                record.author or '',
                record.region or '',
                record.native_habitat or '',
                record.bloom_time or '',
                record.growth_habit or '',
                record.climate_preference or '',
                record.light_requirements or '',
                record.temperature_range or '',
                record.image_filename or '',
                record.photographer or '',
                record.image_source or '',
                record.ai_description or '',
                record.ai_confidence or '',
                record.ingestion_source or '',
                record.validation_status or '',
                'Yes' if record.is_featured else 'No',
                record.view_count or 0,
                record.created_at.strftime('%Y-%m-%d %H:%M') if record.created_at else '',
                record.updated_at.strftime('%Y-%m-%d %H:%M') if record.updated_at else ''
            ]
            
            if include_judging:
                # Get latest judging analysis
                latest_judging = JudgingAnalysis.query.filter_by(
                    orchid_id=record.id
                ).order_by(JudgingAnalysis.analysis_date.desc()).first()
                
                if latest_judging:
                    row_data.extend([
                        f"{latest_judging.score:.1f} ({latest_judging.percentage:.1f}%)",
                        latest_judging.suggested_award_level or '',
                        'Yes' if latest_judging.is_award_worthy else 'No',
                        latest_judging.analysis_date.strftime('%Y-%m-%d'),
                        latest_judging.judging_standard.organization if latest_judging.judging_standard else ''
                    ])
                else:
                    row_data.extend(['', '', '', '', ''])
            
            # Write row
            for col_num, value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Add metadata sheet
        self._add_metadata_sheet(wb, len(records), user_id)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def export_to_csv(self, records: List[OrchidRecord], user_id: Optional[int] = None,
                     include_judging: bool = False) -> StringIO:
        """
        Export orchid records to CSV format
        
        Args:
            records: List of OrchidRecord objects
            user_id: User ID for filtering (None for admin export)
            include_judging: Whether to include judging analysis
            
        Returns:
            StringIO object with CSV data
        """
        output = StringIO()
        
        # Define headers
        headers = [
            'ID', 'User_ID', 'Plant_ID', 'Display_Name', 'Scientific_Name',
            'Genus', 'Species', 'Author', 'Region', 'Native_Habitat',
            'Bloom_Time', 'Growth_Habit', 'Climate_Preference',
            'Light_Requirements', 'Temperature_Range',
            'Image_Filename', 'Photographer', 'Source',
            'AI_Description', 'AI_Confidence', 'Ingestion_Source',
            'Validation_Status', 'Featured', 'View_Count',
            'Created_Date', 'Updated_Date'
        ]
        
        if include_judging:
            headers.extend([
                'Judging_Score', 'Judging_Percentage', 'Award_Level',
                'Award_Worthy', 'Last_Judged', 'Judging_Organization'
            ])
        
        writer = csv.writer(output)
        writer.writerow(headers)
        
        # Add data rows
        for record in records:
            # Get associated upload info
            upload = UserUpload.query.filter_by(orchid_id=record.id).first()
            user = User.query.get(record.user_id) if record.user_id else None
            
            row_data = [
                record.id,
                user.user_id if user else '',
                upload.plant_id if upload else '',
                record.display_name or '',
                record.scientific_name or '',
                record.genus or '',
                record.species or '',
                record.author or '',
                record.region or '',
                record.native_habitat or '',
                record.bloom_time or '',
                record.growth_habit or '',
                record.climate_preference or '',
                record.light_requirements or '',
                record.temperature_range or '',
                record.image_filename or '',
                record.photographer or '',
                record.image_source or '',
                record.ai_description or '',
                record.ai_confidence or '',
                record.ingestion_source or '',
                record.validation_status or '',
                'Yes' if record.is_featured else 'No',
                record.view_count or 0,
                record.created_at.strftime('%Y-%m-%d %H:%M') if record.created_at else '',
                record.updated_at.strftime('%Y-%m-%d %H:%M') if record.updated_at else ''
            ]
            
            if include_judging:
                # Get latest judging analysis
                latest_judging = JudgingAnalysis.query.filter_by(
                    orchid_id=record.id
                ).order_by(JudgingAnalysis.analysis_date.desc()).first()
                
                if latest_judging:
                    row_data.extend([
                        latest_judging.score or '',
                        latest_judging.percentage or '',
                        latest_judging.suggested_award_level or '',
                        'Yes' if latest_judging.is_award_worthy else 'No',
                        latest_judging.analysis_date.strftime('%Y-%m-%d') if latest_judging.analysis_date else '',
                        latest_judging.judging_standard.organization if latest_judging.judging_standard else ''
                    ])
                else:
                    row_data.extend(['', '', '', '', '', ''])
            
            writer.writerow(row_data)
        
        output.seek(0)
        return output
    
    def export_user_data(self, user_id: int, format: str = 'excel') -> Any:
        """
        Export individual user's data
        
        Args:
            user_id: User ID to export data for
            format: 'excel' or 'csv'
            
        Returns:
            BytesIO (Excel) or StringIO (CSV) with user's data
        """
        # Get user's orchid records
        records = OrchidRecord.query.filter_by(user_id=user_id).all()
        
        if format.lower() == 'csv':
            return self.export_to_csv(records, user_id, include_judging=True)
        else:
            return self.export_to_excel(records, user_id, include_judging=True)
    
    def export_judging_analysis(self, orchid_id: int) -> BytesIO:
        """
        Export detailed judging analysis for a specific orchid
        
        Args:
            orchid_id: Orchid ID to export analysis for
            
        Returns:
            BytesIO object with Excel data
        """
        orchid = OrchidRecord.query.get_or_404(orchid_id)
        analyses = JudgingAnalysis.query.filter_by(orchid_id=orchid_id).order_by(
            JudgingAnalysis.analysis_date.desc()
        ).all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Judging Analysis"
        
        # Add orchid information
        ws['A1'] = "Orchid Information"
        ws['A1'].font = Font(bold=True, size=14)
        
        info_data = [
            ('Orchid ID:', orchid.id),
            ('Scientific Name:', orchid.scientific_name or 'Unknown'),
            ('Display Name:', orchid.display_name or ''),
            ('Genus:', orchid.genus or ''),
            ('Species:', orchid.species or ''),
        ]
        
        for row_num, (label, value) in enumerate(info_data, start=2):
            ws[f'A{row_num}'] = label
            ws[f'B{row_num}'] = value
            ws[f'A{row_num}'].font = Font(bold=True)
        
        # Add judging analyses
        start_row = len(info_data) + 4
        ws[f'A{start_row}'] = "Judging Analyses"
        ws[f'A{start_row}'].font = Font(bold=True, size=14)
        
        # Headers for analyses
        headers = ['Date', 'Organization', 'Category', 'Score', 'Percentage', 
                  'Award Level', 'Award Worthy', 'Confidence', 'Comments']
        
        header_row = start_row + 2
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # Add analysis data
        for row_num, analysis in enumerate(analyses, start=header_row + 1):
            row_data = [
                analysis.analysis_date.strftime('%Y-%m-%d %H:%M') if analysis.analysis_date else '',
                analysis.judging_standard.organization if analysis.judging_standard else '',
                analysis.judging_standard.category if analysis.judging_standard else '',
                f"{analysis.score:.1f}" if analysis.score else '',
                f"{analysis.percentage:.1f}%" if analysis.percentage else '',
                analysis.suggested_award_level or '',
                'Yes' if analysis.is_award_worthy else 'No',
                f"{analysis.confidence_level:.2f}" if analysis.confidence_level else '',
                analysis.ai_comments or ''
            ]
            
            for col_num, value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _style_headers(self, worksheet, headers):
        """Apply styling to header row"""
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_metadata_sheet(self, workbook, record_count, user_id):
        """Add metadata sheet with export information"""
        ws = workbook.create_sheet("Export Metadata")
        
        metadata = [
            ('Export Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Record Count:', record_count),
            ('Export Type:', 'User Export' if user_id else 'Admin Export'),
            ('Exported By:', f'User ID: {user_id}' if user_id else 'Administrator'),
            ('System:', 'Orchid Continuum'),
            ('Version:', '1.0'),
            ('Disclaimer:', 'This data is for educational and personal use only.'),
            ('Note:', 'Judging analyses are for entertainment purposes and not official awards.')
        ]
        
        for row_num, (label, value) in enumerate(metadata, start=1):
            ws[f'A{row_num}'] = label
            ws[f'B{row_num}'] = value
            ws[f'A{row_num}'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

# Global exporter instance
exporter = OrchidDataExporter()

def export_orchid_data(record_ids: List[int], format: str = 'excel', 
                      user_id: Optional[int] = None, include_judging: bool = False) -> Any:
    """
    Convenience function to export orchid data
    
    Args:
        record_ids: List of orchid record IDs to export
        format: 'excel' or 'csv'
        user_id: User ID for permission checking
        include_judging: Include judging analysis data
        
    Returns:
        BytesIO or StringIO with export data
    """
    records = OrchidRecord.query.filter(OrchidRecord.id.in_(record_ids)).all()
    
    if format.lower() == 'csv':
        return exporter.export_to_csv(records, user_id, include_judging)
    else:
        return exporter.export_to_excel(records, user_id, include_judging=include_judging)

def get_export_filename(format: str, user_id: Optional[int] = None, 
                       timestamp: bool = True) -> str:
    """
    Generate appropriate filename for export
    
    Args:
        format: File format ('excel' or 'csv')
        user_id: User ID (for user-specific exports)
        timestamp: Include timestamp in filename
        
    Returns:
        Generated filename
    """
    base_name = 'orchid_data'
    
    if user_id:
        base_name += f'_user_{user_id}'
    else:
        base_name += '_full'
    
    if timestamp:
        timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_name += f'_{timestamp_str}'
    
    extension = 'xlsx' if format.lower() == 'excel' else 'csv'
    return f'{base_name}.{extension}'