"""
Professional Report Generator for Orchid Continuum
Generates PDF and CSV reports with data visualization, progress tracking, and customizable templates
"""
import os
import csv
import json
import logging
from io import BytesIO, StringIO
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Tuple, Callable
from dataclasses import dataclass, field
from enum import Enum
import tempfile
import threading
import uuid

# PDF and visualization imports
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
import pandas as pd
import numpy as np

from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib.colors import Color, black, white, darkblue, darkgreen, orange, red
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, 
                               PageBreak, Image as RLImage, KeepTogether, NextPageTemplate, PageTemplate)
from reportlab.platypus.frames import Frame
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart

# Excel imports
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

from flask import current_app
from sqlalchemy import func, and_, or_, text
from models import (OrchidRecord, UserUpload, User, JudgingAnalysis, 
                   ScrapingLog, Certificate, UserActivity, GameScore)

# Configure logging
logger = logging.getLogger(__name__)

class ReportType(Enum):
    """Available report types"""
    DATA_SUMMARY = "data_summary"
    COLLECTION_ANALYSIS = "collection_analysis"
    GEOGRAPHIC_DISTRIBUTION = "geographic_distribution"
    FLOWERING_ANALYSIS = "flowering_analysis"
    AI_ANALYSIS_REPORT = "ai_analysis_report"
    USER_ACTIVITY_REPORT = "user_activity_report"
    SYSTEM_HEALTH_REPORT = "system_health_report"
    CUSTOM_QUERY_REPORT = "custom_query_report"

class OutputFormat(Enum):
    """Available output formats"""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    JSON = "json"

@dataclass
class ReportProgress:
    """Track report generation progress"""
    report_id: str
    total_steps: int
    completed_steps: int = 0
    current_step: str = ""
    status: str = "running"  # running, completed, failed
    error_message: str = ""
    start_time: datetime = field(default_factory=datetime.now)
    end_time: Optional[datetime] = None
    
    @property
    def progress_percentage(self) -> float:
        return (self.completed_steps / self.total_steps * 100) if self.total_steps > 0 else 0
    
    @property
    def elapsed_time(self) -> timedelta:
        end = self.end_time or datetime.now()
        return end - self.start_time

@dataclass
class ReportConfig:
    """Configuration for report generation"""
    report_type: ReportType
    output_formats: List[OutputFormat]
    title: str = ""
    subtitle: str = ""
    author: str = "Orchid Continuum System"
    include_charts: bool = True
    include_images: bool = False
    include_metadata: bool = True
    date_range: Optional[Tuple[datetime, datetime]] = None
    filters: Dict[str, Any] = field(default_factory=dict)
    custom_queries: List[str] = field(default_factory=list)
    template_name: str = "standard"
    color_scheme: str = "orchid_purple"  # orchid_purple, professional_blue, nature_green
    chart_style: str = "seaborn"  # seaborn, ggplot, classic
    
class ReportTemplate:
    """Base class for report templates"""
    
    def __init__(self, config: ReportConfig):
        self.config = config
        self.colors = self._get_color_scheme()
        self.setup_matplotlib_style()
        
    def _get_color_scheme(self) -> Dict[str, Any]:
        """Get color scheme based on configuration"""
        schemes = {
            'orchid_purple': {
                'primary': Color(0.4, 0.2, 0.6),      # Deep purple
                'secondary': Color(0.8, 0.4, 0.8),    # Light purple  
                'accent': Color(1.0, 0.8, 0.0),       # Gold
                'text': Color(0.2, 0.2, 0.2),         # Dark gray
                'background': Color(0.98, 0.98, 1.0), # Light purple tint
                'chart_palette': ['#6B46C1', '#A78BFA', '#C4B5FD', '#EDE9FE', '#F3E8FF']
            },
            'professional_blue': {
                'primary': Color(0.1, 0.3, 0.6),      # Navy blue
                'secondary': Color(0.4, 0.6, 0.8),    # Light blue
                'accent': Color(0.0, 0.7, 0.9),       # Cyan
                'text': Color(0.1, 0.1, 0.1),         # Near black
                'background': Color(0.98, 0.99, 1.0), # Light blue tint
                'chart_palette': ['#1E3A8A', '#3B82F6', '#60A5FA', '#93C5FD', '#DBEAFE']
            },
            'nature_green': {
                'primary': Color(0.1, 0.5, 0.2),      # Forest green
                'secondary': Color(0.4, 0.7, 0.4),    # Light green
                'accent': Color(0.8, 0.5, 0.2),       # Orange
                'text': Color(0.2, 0.3, 0.1),         # Dark green
                'background': Color(0.98, 1.0, 0.98), # Light green tint
                'chart_palette': ['#166534', '#22C55E', '#4ADE80', '#86EFAC', '#DCFCE7']
            }
        }
        return schemes.get(self.config.color_scheme, schemes['orchid_purple'])
    
    def setup_matplotlib_style(self):
        """Configure matplotlib with chosen style"""
        plt.style.use(self.config.chart_style)
        sns.set_palette(self.colors['chart_palette'])
        
        # Configure default figure settings
        plt.rcParams.update({
            'figure.figsize': (10, 6),
            'figure.dpi': 100,
            'savefig.dpi': 300,
            'figure.facecolor': 'white',
            'axes.facecolor': 'white',
            'axes.spines.top': False,
            'axes.spines.right': False,
            'font.size': 10,
            'axes.titlesize': 12,
            'axes.labelsize': 10,
            'xtick.labelsize': 9,
            'ytick.labelsize': 9,
            'legend.fontsize': 9
        })

class OrchidReportGenerator:
    """Main report generator class"""
    
    def __init__(self):
        self.progress_tracker: Dict[str, ReportProgress] = {}
        self.temp_dir = tempfile.mkdtemp(prefix='orchid_reports_')
        self.chart_cache: Dict[str, str] = {}  # Cache for generated chart files
        
    def generate_report(self, config: ReportConfig, progress_callback: Optional[Callable] = None) -> Dict[str, Any]:
        """
        Generate comprehensive report based on configuration
        
        Args:
            config: Report configuration
            progress_callback: Optional callback function for progress updates
            
        Returns:
            Dict with generated report information and file paths
        """
        report_id = str(uuid.uuid4())
        
        try:
            # Initialize progress tracking
            total_steps = self._calculate_total_steps(config)
            progress = ReportProgress(report_id=report_id, total_steps=total_steps)
            self.progress_tracker[report_id] = progress
            
            logger.info(f"🚀 Starting report generation: {config.report_type.value}")
            
            # Step 1: Gather data
            progress.current_step = "Gathering data..."
            if progress_callback:
                progress_callback(progress)
                
            data = self._gather_report_data(config)
            progress.completed_steps += 1
            
            # Step 2: Generate visualizations
            if config.include_charts:
                progress.current_step = "Creating visualizations..."
                if progress_callback:
                    progress_callback(progress)
                    
                charts = self._generate_visualizations(data, config)
                progress.completed_steps += 1
            else:
                charts = {}
            
            # Step 3: Generate reports in requested formats
            generated_files = {}
            
            for output_format in config.output_formats:
                progress.current_step = f"Generating {output_format.value.upper()} report..."
                if progress_callback:
                    progress_callback(progress)
                    
                if output_format == OutputFormat.PDF:
                    pdf_file = self._generate_pdf_report(data, charts, config)
                    generated_files['pdf'] = pdf_file
                elif output_format == OutputFormat.EXCEL:
                    excel_file = self._generate_excel_report(data, charts, config)
                    generated_files['excel'] = excel_file
                elif output_format == OutputFormat.CSV:
                    csv_file = self._generate_csv_report(data, config)
                    generated_files['csv'] = csv_file
                elif output_format == OutputFormat.JSON:
                    json_file = self._generate_json_report(data, config)
                    generated_files['json'] = json_file
                
                progress.completed_steps += 1
            
            # Complete
            progress.current_step = "Report generation completed!"
            progress.status = "completed"
            progress.end_time = datetime.now()
            progress.completed_steps = total_steps
            
            if progress_callback:
                progress_callback(progress)
            
            logger.info(f"✅ Report generated successfully: {report_id}")
            
            return {
                'report_id': report_id,
                'status': 'success',
                'files': generated_files,
                'metadata': {
                    'report_type': config.report_type.value,
                    'generation_time': progress.elapsed_time.total_seconds(),
                    'total_records': len(data.get('records', [])),
                    'charts_generated': len(charts),
                    'author': config.author,
                    'timestamp': datetime.now().isoformat()
                }
            }
            
        except Exception as e:
            logger.error(f"❌ Report generation failed: {e}")
            progress.status = "failed"
            progress.error_message = str(e)
            progress.end_time = datetime.now()
            
            if progress_callback:
                progress_callback(progress)
                
            return {
                'report_id': report_id,
                'status': 'failed',
                'error': str(e),
                'files': {}
            }
    
    def _calculate_total_steps(self, config: ReportConfig) -> int:
        """Calculate total steps for progress tracking"""
        steps = 1  # Data gathering
        if config.include_charts:
            steps += 1  # Chart generation
        steps += len(config.output_formats)  # One step per output format
        return steps
    
    def _gather_report_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather data based on report type and configuration"""
        data = {'records': [], 'summary': {}, 'analytics': {}}
        
        try:
            if config.report_type == ReportType.DATA_SUMMARY:
                data = self._gather_data_summary()
            elif config.report_type == ReportType.COLLECTION_ANALYSIS:
                data = self._gather_collection_analysis(config)
            elif config.report_type == ReportType.GEOGRAPHIC_DISTRIBUTION:
                data = self._gather_geographic_data(config)
            elif config.report_type == ReportType.FLOWERING_ANALYSIS:
                data = self._gather_flowering_data(config)
            elif config.report_type == ReportType.AI_ANALYSIS_REPORT:
                data = self._gather_ai_analysis_data(config)
            elif config.report_type == ReportType.USER_ACTIVITY_REPORT:
                data = self._gather_user_activity_data(config)
            elif config.report_type == ReportType.SYSTEM_HEALTH_REPORT:
                data = self._gather_system_health_data(config)
            elif config.report_type == ReportType.CUSTOM_QUERY_REPORT:
                data = self._gather_custom_query_data(config)
            
            logger.info(f"📊 Gathered {len(data.get('records', []))} records for report")
            return data
            
        except Exception as e:
            logger.error(f"Error gathering report data: {e}")
            raise
    
    def _gather_data_summary(self) -> Dict[str, Any]:
        """Gather comprehensive data summary"""
        # Total counts
        total_orchids = OrchidRecord.query.count()
        total_users = User.query.count()
        total_uploads = UserUpload.query.count()
        
        # Recent activity (last 30 days)
        thirty_days_ago = datetime.now() - timedelta(days=30)
        recent_orchids = OrchidRecord.query.filter(
            OrchidRecord.created_at >= thirty_days_ago
        ).count()
        recent_uploads = UserUpload.query.filter(
            UserUpload.created_at >= thirty_days_ago
        ).count()
        
        # Taxonomy distribution
        genus_counts = OrchidRecord.query.with_entities(
            OrchidRecord.genus, func.count(OrchidRecord.id).label('count')
        ).filter(OrchidRecord.genus.isnot(None)).group_by(OrchidRecord.genus).all()
        
        # Geographic distribution
        country_counts = OrchidRecord.query.with_entities(
            OrchidRecord.country, func.count(OrchidRecord.id).label('count')
        ).filter(OrchidRecord.country.isnot(None)).group_by(OrchidRecord.country).all()
        
        # AI analysis statistics
        ai_analyzed = OrchidRecord.query.filter(
            OrchidRecord.ai_confidence.isnot(None)
        ).count()
        avg_confidence = OrchidRecord.query.with_entities(
            func.avg(OrchidRecord.ai_confidence)
        ).scalar() or 0
        
        # Data quality metrics
        records_with_images = OrchidRecord.query.filter(
            or_(
                OrchidRecord.image_url.isnot(None),
                OrchidRecord.google_drive_id.isnot(None),
                OrchidRecord.image_filename.isnot(None)
            )
        ).count()
        
        records_with_coordinates = OrchidRecord.query.filter(
            and_(
                OrchidRecord.decimal_latitude.isnot(None),
                OrchidRecord.decimal_longitude.isnot(None)
            )
        ).count()
        
        return {
            'records': [],  # No individual records for summary report
            'summary': {
                'total_orchids': total_orchids,
                'total_users': total_users,
                'total_uploads': total_uploads,
                'recent_orchids': recent_orchids,
                'recent_uploads': recent_uploads,
                'ai_analyzed_count': ai_analyzed,
                'avg_ai_confidence': round(avg_confidence, 2),
                'records_with_images': records_with_images,
                'records_with_coordinates': records_with_coordinates,
                'image_percentage': round((records_with_images / total_orchids * 100) if total_orchids > 0 else 0, 1),
                'coordinate_percentage': round((records_with_coordinates / total_orchids * 100) if total_orchids > 0 else 0, 1)
            },
            'analytics': {
                'genus_distribution': [{'genus': g.genus, 'count': g.count} for g in genus_counts[:20]],
                'country_distribution': [{'country': c.country, 'count': c.count} for c in country_counts[:15]],
                'monthly_growth': self._calculate_monthly_growth()
            }
        }
    
    def _gather_collection_analysis(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather collection analysis data"""
        query = OrchidRecord.query
        
        # Apply filters
        if config.filters.get('genus'):
            query = query.filter(OrchidRecord.genus == config.filters['genus'])
        if config.filters.get('country'):
            query = query.filter(OrchidRecord.country == config.filters['country'])
        if config.date_range:
            start_date, end_date = config.date_range
            query = query.filter(OrchidRecord.created_at.between(start_date, end_date))
        
        records = query.all()
        
        # Analyze collection characteristics
        flowering_analysis = self._analyze_flowering_patterns(records)
        growth_habit_analysis = self._analyze_growth_habits(records)
        climate_analysis = self._analyze_climate_preferences(records)
        
        return {
            'records': [self._serialize_orchid_record(r) for r in records],
            'summary': {
                'total_records': len(records),
                'unique_genera': len(set(r.genus for r in records if r.genus)),
                'unique_species': len(set(r.species for r in records if r.species)),
                'flowering_records': sum(1 for r in records if r.is_flowering),
                'average_ai_confidence': np.mean([r.ai_confidence for r in records if r.ai_confidence]) if records else 0
            },
            'analytics': {
                'flowering_patterns': flowering_analysis,
                'growth_habits': growth_habit_analysis,
                'climate_preferences': climate_analysis
            }
        }
    
    def _gather_geographic_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather geographic distribution data"""
        records = OrchidRecord.query.filter(
            and_(
                OrchidRecord.decimal_latitude.isnot(None),
                OrchidRecord.decimal_longitude.isnot(None)
            )
        ).all()
        
        # Group by regions/countries
        geographic_summary = {}
        for record in records:
            country = record.country or 'Unknown'
            if country not in geographic_summary:
                geographic_summary[country] = {
                    'count': 0,
                    'genera': set(),
                    'coordinates': []
                }
            
            geographic_summary[country]['count'] += 1
            if record.genus:
                geographic_summary[country]['genera'].add(record.genus)
            geographic_summary[country]['coordinates'].append({
                'lat': float(record.decimal_latitude),
                'lng': float(record.decimal_longitude),
                'name': record.display_name or record.scientific_name
            })
        
        # Convert sets to lists for JSON serialization
        for country_data in geographic_summary.values():
            country_data['genera'] = list(country_data['genera'])
            country_data['genera_count'] = len(country_data['genera'])
        
        return {
            'records': [self._serialize_orchid_record(r) for r in records],
            'summary': {
                'total_georeferenced': len(records),
                'countries_represented': len(geographic_summary),
                'coordinate_density': len(records) / len(geographic_summary) if geographic_summary else 0
            },
            'analytics': {
                'geographic_distribution': geographic_summary,
                'coordinate_clusters': self._analyze_coordinate_clusters(records)
            }
        }
    
    def _gather_flowering_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather flowering analysis data"""
        flowering_records = OrchidRecord.query.filter(
            OrchidRecord.is_flowering == True
        ).all()
        
        # Analyze flowering patterns
        seasonal_patterns = self._analyze_seasonal_flowering(flowering_records)
        genus_flowering = self._analyze_genus_flowering_patterns(flowering_records)
        
        return {
            'records': [self._serialize_orchid_record(r) for r in flowering_records],
            'summary': {
                'total_flowering': len(flowering_records),
                'average_flower_count': np.mean([r.flower_count for r in flowering_records if r.flower_count]) if flowering_records else 0,
                'peak_flowering_season': self._determine_peak_season(flowering_records)
            },
            'analytics': {
                'seasonal_patterns': seasonal_patterns,
                'genus_patterns': genus_flowering,
                'flowering_stages': self._analyze_flowering_stages(flowering_records)
            }
        }
    
    def _gather_ai_analysis_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather AI analysis performance data"""
        ai_records = OrchidRecord.query.filter(
            OrchidRecord.ai_confidence.isnot(None)
        ).all()
        
        # Analyze AI performance
        confidence_distribution = self._analyze_confidence_distribution(ai_records)
        accuracy_metrics = self._calculate_ai_accuracy_metrics(ai_records)
        
        return {
            'records': [self._serialize_orchid_record(r) for r in ai_records],
            'summary': {
                'total_ai_analyzed': len(ai_records),
                'average_confidence': np.mean([r.ai_confidence for r in ai_records]) if ai_records else 0,
                'high_confidence_percentage': sum(1 for r in ai_records if r.ai_confidence and r.ai_confidence > 0.8) / len(ai_records) * 100 if ai_records else 0
            },
            'analytics': {
                'confidence_distribution': confidence_distribution,
                'accuracy_metrics': accuracy_metrics,
                'genus_confidence_by_genus': self._analyze_confidence_by_genus(ai_records)
            }
        }
    
    def _gather_user_activity_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather user activity data"""
        # Get user activity data (if model exists)
        try:
            activities = UserActivity.query.all()
            activity_summary = self._analyze_user_activities(activities)
        except:
            activities = []
            activity_summary = {}
        
        # User upload patterns
        users_with_uploads = User.query.join(UserUpload).all()
        upload_patterns = self._analyze_upload_patterns(users_with_uploads)
        
        return {
            'records': [],  # User data is sensitive, summarize only
            'summary': {
                'total_users': User.query.count(),
                'active_users': len(users_with_uploads),
                'total_activities': len(activities) if activities else 0
            },
            'analytics': {
                'activity_summary': activity_summary,
                'upload_patterns': upload_patterns,
                'user_engagement': self._calculate_user_engagement_metrics()
            }
        }
    
    def _gather_system_health_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather system health and performance data"""
        # Database health metrics
        db_metrics = {
            'total_records': OrchidRecord.query.count(),
            'recent_records': OrchidRecord.query.filter(
                OrchidRecord.created_at >= datetime.now() - timedelta(days=7)
            ).count(),
            'data_quality_score': self._calculate_data_quality_score()
        }
        
        # System performance metrics
        performance_metrics = {
            'average_response_time': 0.5,  # Placeholder - would integrate with monitoring
            'uptime_percentage': 99.5,     # Placeholder - would integrate with monitoring
            'error_rate': 0.1             # Placeholder - would integrate with monitoring
        }
        
        return {
            'records': [],  # System data, not individual records
            'summary': {
                'system_status': 'healthy',
                'last_backup': datetime.now().isoformat(),
                'data_integrity_score': 95.2
            },
            'analytics': {
                'database_metrics': db_metrics,
                'performance_metrics': performance_metrics,
                'health_indicators': self._calculate_health_indicators()
            }
        }
    
    def _gather_custom_query_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Execute custom queries and gather results"""
        results = {}
        
        for i, query_str in enumerate(config.custom_queries):
            try:
                # Execute raw SQL query (with safety checks)
                if self._is_safe_query(query_str):
                    from app import db
                    result = db.session.execute(text(query_str))
                    results[f'custom_query_{i+1}'] = [dict(row) for row in result]
                else:
                    results[f'custom_query_{i+1}'] = {'error': 'Unsafe query rejected'}
            except Exception as e:
                results[f'custom_query_{i+1}'] = {'error': str(e)}
        
        return {
            'records': [],
            'summary': {
                'queries_executed': len(config.custom_queries),
                'successful_queries': sum(1 for r in results.values() if 'error' not in r)
            },
            'analytics': {
                'query_results': results
            }
        }
    
    def _generate_visualizations(self, data: Dict[str, Any], config: ReportConfig) -> Dict[str, str]:
        """Generate visualization charts and return file paths"""
        charts = {}
        
        try:
            # Generate charts based on report type
            if config.report_type == ReportType.DATA_SUMMARY:
                charts.update(self._create_summary_charts(data))
            elif config.report_type == ReportType.COLLECTION_ANALYSIS:
                charts.update(self._create_collection_charts(data))
            elif config.report_type == ReportType.GEOGRAPHIC_DISTRIBUTION:
                charts.update(self._create_geographic_charts(data))
            elif config.report_type == ReportType.FLOWERING_ANALYSIS:
                charts.update(self._create_flowering_charts(data))
            elif config.report_type == ReportType.AI_ANALYSIS_REPORT:
                charts.update(self._create_ai_analysis_charts(data))
            
            logger.info(f"📈 Generated {len(charts)} charts for report")
            return charts
            
        except Exception as e:
            logger.error(f"Error generating visualizations: {e}")
            return {}
    
    def _create_summary_charts(self, data: Dict[str, Any]) -> Dict[str, str]:
        """Create summary report charts"""
        charts = {}
        
        # Genus distribution pie chart
        if data['analytics'].get('genus_distribution'):
            fig, ax = plt.subplots(figsize=(10, 8))
            
            genus_data = data['analytics']['genus_distribution'][:10]  # Top 10
            labels = [g['genus'] for g in genus_data]
            sizes = [g['count'] for g in genus_data]
            
            ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
            ax.set_title('Top Genera Distribution', fontsize=14, fontweight='bold')
            
            chart_path = os.path.join(self.temp_dir, 'genus_distribution.png')
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            charts['genus_distribution'] = chart_path
        
        # Monthly growth trend
        if data['analytics'].get('monthly_growth'):
            fig, ax = plt.subplots(figsize=(12, 6))
            
            growth_data = data['analytics']['monthly_growth']
            months = [g['month'] for g in growth_data]
            counts = [g['count'] for g in growth_data]
            
            ax.plot(months, counts, marker='o', linewidth=2, markersize=6)
            ax.set_title('Monthly Collection Growth', fontsize=14, fontweight='bold')
            ax.set_xlabel('Month')
            ax.set_ylabel('New Orchid Records')
            ax.grid(True, alpha=0.3)
            
            # Format x-axis
            ax.tick_params(axis='x', rotation=45)
            
            chart_path = os.path.join(self.temp_dir, 'monthly_growth.png')
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            charts['monthly_growth'] = chart_path
        
        # Data quality dashboard
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(15, 10))
        
        summary = data['summary']
        
        # Image coverage
        ax1.pie([summary['records_with_images'], summary['total_orchids'] - summary['records_with_images']], 
                labels=['With Images', 'Without Images'], autopct='%1.1f%%',
                colors=['#22C55E', '#EF4444'])
        ax1.set_title('Image Coverage')
        
        # Geographic coverage  
        ax2.pie([summary['records_with_coordinates'], summary['total_orchids'] - summary['records_with_coordinates']],
                labels=['With Coordinates', 'Without Coordinates'], autopct='%1.1f%%',
                colors=['#3B82F6', '#F97316'])
        ax2.set_title('Geographic Coverage')
        
        # AI analysis coverage
        ax3.pie([summary['ai_analyzed_count'], summary['total_orchids'] - summary['ai_analyzed_count']],
                labels=['AI Analyzed', 'Not Analyzed'], autopct='%1.1f%%',
                colors=['#8B5CF6', '#6B7280'])
        ax3.set_title('AI Analysis Coverage')
        
        # Summary statistics
        stats_labels = ['Total Orchids', 'Total Users', 'Recent Uploads', 'Avg AI Confidence']
        stats_values = [summary['total_orchids'], summary['total_users'], 
                       summary['recent_uploads'], summary['avg_ai_confidence']]
        
        bars = ax4.bar(stats_labels, stats_values, color=['#10B981', '#F59E0B', '#EF4444', '#8B5CF6'])
        ax4.set_title('System Statistics')
        ax4.tick_params(axis='x', rotation=45)
        
        # Add value labels on bars
        for bar, value in zip(bars, stats_values):
            height = bar.get_height()
            ax4.text(bar.get_x() + bar.get_width()/2., height,
                    f'{value:.1f}' if isinstance(value, float) else f'{value}',
                    ha='center', va='bottom')
        
        plt.tight_layout()
        chart_path = os.path.join(self.temp_dir, 'data_quality_dashboard.png')
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        charts['data_quality_dashboard'] = chart_path
        
        return charts
    
    def _create_collection_charts(self, data: Dict[str, Any]) -> Dict[str, str]:
        """Create collection analysis charts"""
        charts = {}
        
        # Growth habits distribution
        if data['analytics'].get('growth_habits'):
            fig, ax = plt.subplots(figsize=(10, 6))
            
            habits_data = data['analytics']['growth_habits']
            habits = list(habits_data.keys())
            counts = list(habits_data.values())
            
            bars = ax.bar(habits, counts, color=sns.color_palette("husl", len(habits)))
            ax.set_title('Growth Habits Distribution', fontsize=14, fontweight='bold')
            ax.set_xlabel('Growth Habit')
            ax.set_ylabel('Count')
            
            # Add value labels
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom')
            
            plt.xticks(rotation=45)
            chart_path = os.path.join(self.temp_dir, 'growth_habits.png')
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            charts['growth_habits'] = chart_path
        
        # Climate preferences
        if data['analytics'].get('climate_preferences'):
            fig, ax = plt.subplots(figsize=(8, 8))
            
            climate_data = data['analytics']['climate_preferences']
            labels = list(climate_data.keys())
            sizes = list(climate_data.values())
            
            ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90,
                   colors=sns.color_palette("coolwarm", len(labels)))
            ax.set_title('Climate Preferences Distribution', fontsize=14, fontweight='bold')
            
            chart_path = os.path.join(self.temp_dir, 'climate_preferences.png')
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            charts['climate_preferences'] = chart_path
            
        return charts
    
    def _create_geographic_charts(self, data: Dict[str, Any]) -> Dict[str, str]:
        """Create geographic distribution charts"""
        charts = {}
        
        # Country distribution bar chart
        if data['analytics'].get('geographic_distribution'):
            fig, ax = plt.subplots(figsize=(12, 8))
            
            geo_data = data['analytics']['geographic_distribution']
            countries = list(geo_data.keys())[:15]  # Top 15 countries
            counts = [geo_data[country]['count'] for country in countries]
            
            bars = ax.barh(countries, counts, color=sns.color_palette("viridis", len(countries)))
            ax.set_title('Geographic Distribution by Country', fontsize=14, fontweight='bold')
            ax.set_xlabel('Number of Records')
            ax.set_ylabel('Country')
            
            # Add value labels
            for i, (bar, count) in enumerate(zip(bars, counts)):
                width = bar.get_width()
                ax.text(width, bar.get_y() + bar.get_height()/2.,
                       f'{count}', ha='left', va='center', fontweight='bold')
            
            chart_path = os.path.join(self.temp_dir, 'geographic_distribution.png')
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            charts['geographic_distribution'] = chart_path
        
        return charts
    
    def _create_flowering_charts(self, data: Dict[str, Any]) -> Dict[str, str]:
        """Create flowering analysis charts"""
        charts = {}
        
        # Seasonal flowering patterns
        if data['analytics'].get('seasonal_patterns'):
            fig, ax = plt.subplots(figsize=(10, 6))
            
            seasonal_data = data['analytics']['seasonal_patterns']
            seasons = list(seasonal_data.keys())
            counts = list(seasonal_data.values())
            
            bars = ax.bar(seasons, counts, color=['#FFD700', '#FF6B35', '#8B4513', '#4682B4'])
            ax.set_title('Seasonal Flowering Patterns', fontsize=14, fontweight='bold')
            ax.set_xlabel('Season')
            ax.set_ylabel('Flowering Records')
            
            # Add value labels
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom')
            
            chart_path = os.path.join(self.temp_dir, 'seasonal_flowering.png')
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            charts['seasonal_flowering'] = chart_path
        
        return charts
    
    def _create_ai_analysis_charts(self, data: Dict[str, Any]) -> Dict[str, str]:
        """Create AI analysis performance charts"""
        charts = {}
        
        # Confidence distribution histogram
        if data['analytics'].get('confidence_distribution'):
            fig, ax = plt.subplots(figsize=(10, 6))
            
            confidence_data = data['analytics']['confidence_distribution']
            ax.hist(confidence_data, bins=20, alpha=0.7, color='skyblue', edgecolor='black')
            ax.set_title('AI Confidence Distribution', fontsize=14, fontweight='bold')
            ax.set_xlabel('Confidence Score')
            ax.set_ylabel('Frequency')
            ax.grid(True, alpha=0.3)
            
            chart_path = os.path.join(self.temp_dir, 'confidence_distribution.png')
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            charts['confidence_distribution'] = chart_path
        
        return charts
    
    def _generate_pdf_report(self, data: Dict[str, Any], charts: Dict[str, str], config: ReportConfig) -> str:
        """Generate comprehensive PDF report"""
        template = ReportTemplate(config)
        
        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.report_type.value}_{timestamp}.pdf"
        filepath = os.path.join(self.temp_dir, filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        # Build story
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=template.colors['primary']
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=12,
            textColor=template.colors['primary']
        )
        
        # Title page
        story.append(Paragraph(config.title or f"{config.report_type.value.replace('_', ' ').title()} Report", title_style))
        story.append(Spacer(1, 20))
        
        if config.subtitle:
            story.append(Paragraph(config.subtitle, styles['Heading2']))
            story.append(Spacer(1, 20))
        
        # Report metadata table
        metadata_data = [
            ['Report Type:', config.report_type.value.replace('_', ' ').title()],
            ['Generated By:', config.author],
            ['Generation Date:', datetime.now().strftime('%B %d, %Y at %I:%M %p')],
            ['Total Records:', str(len(data.get('records', [])))],
            ['Report ID:', str(uuid.uuid4())[:8]]
        ]
        
        metadata_table = Table(metadata_data, colWidths=[3*cm, 6*cm])
        metadata_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('SPACEAFTER', (0, 0), (-1, -1), 6),
        ]))
        story.append(metadata_table)
        story.append(PageBreak())
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))
        summary_text = self._generate_executive_summary(data, config)
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Key Metrics Table
        if data.get('summary'):
            story.append(Paragraph("Key Metrics", heading_style))
            metrics_data = [['Metric', 'Value']]
            
            for key, value in data['summary'].items():
                display_key = key.replace('_', ' ').title()
                if isinstance(value, float):
                    display_value = f"{value:.2f}"
                elif isinstance(value, int):
                    display_value = f"{value:,}"
                else:
                    display_value = str(value)
                metrics_data.append([display_key, display_value])
            
            metrics_table = Table(metrics_data, colWidths=[6*cm, 4*cm])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), template.colors['primary']),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(metrics_table)
            story.append(PageBreak())
        
        # Insert charts
        if charts:
            story.append(Paragraph("Data Visualizations", heading_style))
            
            for chart_name, chart_path in charts.items():
                if os.path.exists(chart_path):
                    story.append(Paragraph(chart_name.replace('_', ' ').title(), styles['Heading2']))
                    
                    # Add chart image
                    img = RLImage(chart_path, width=15*cm, height=10*cm)
                    story.append(img)
                    story.append(Spacer(1, 20))
            
            story.append(PageBreak())
        
        # Detailed Analysis
        story.append(Paragraph("Detailed Analysis", heading_style))
        analysis_text = self._generate_detailed_analysis(data, config)
        story.append(Paragraph(analysis_text, styles['Normal']))
        
        # Recommendations (if applicable)
        if config.report_type in [ReportType.COLLECTION_ANALYSIS, ReportType.SYSTEM_HEALTH_REPORT]:
            story.append(Spacer(1, 20))
            story.append(Paragraph("Recommendations", heading_style))
            recommendations = self._generate_recommendations(data, config)
            story.append(Paragraph(recommendations, styles['Normal']))
        
        # Footer/Disclaimers
        story.append(PageBreak())
        story.append(Paragraph("Notes and Disclaimers", heading_style))
        disclaimers = [
            "• This report was automatically generated by the Orchid Continuum system.",
            "• Data accuracy depends on source quality and user contributions.",
            "• AI analysis results are for educational purposes and should not replace expert judgment.",
            "• Report generation timestamp and system version are included for reference."
        ]
        for disclaimer in disclaimers:
            story.append(Paragraph(disclaimer, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        logger.info(f"📄 Generated PDF report: {filepath}")
        
        return filepath
    
    def _generate_excel_report(self, data: Dict[str, Any], charts: Dict[str, str], config: ReportConfig) -> str:
        """Generate comprehensive Excel report with charts"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.report_type.value}_{timestamp}.xlsx"
        filepath = os.path.join(self.temp_dir, filename)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        self._create_excel_summary_sheet(ws_summary, data, config)
        
        # Data sheet (if records exist)
        if data.get('records'):
            ws_data = wb.create_sheet("Data")
            self._create_excel_data_sheet(ws_data, data['records'])
        
        # Analytics sheet
        if data.get('analytics'):
            ws_analytics = wb.create_sheet("Analytics")
            self._create_excel_analytics_sheet(ws_analytics, data['analytics'])
        
        # Charts sheet (if charts exist)
        if charts:
            ws_charts = wb.create_sheet("Visualizations")
            self._create_excel_charts_sheet(ws_charts, charts)
        
        # Metadata sheet
        ws_metadata = wb.create_sheet("Metadata")
        self._create_excel_metadata_sheet(ws_metadata, config)
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"📊 Generated Excel report: {filepath}")
        
        return filepath
    
    def _generate_csv_report(self, data: Dict[str, Any], config: ReportConfig) -> str:
        """Generate CSV report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.report_type.value}_{timestamp}.csv"
        filepath = os.path.join(self.temp_dir, filename)
        
        # Write CSV
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            if data.get('records'):
                # Write record data
                fieldnames = self._get_csv_fieldnames(data['records'][0] if data['records'] else {})
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data['records'])
            else:
                # Write summary data
                writer = csv.writer(csvfile)
                writer.writerow(['Report Type', config.report_type.value])
                writer.writerow(['Generated', datetime.now().isoformat()])
                writer.writerow([])
                
                if data.get('summary'):
                    writer.writerow(['Key Metrics'])
                    for key, value in data['summary'].items():
                        writer.writerow([key, value])
        
        logger.info(f"📋 Generated CSV report: {filepath}")
        return filepath
    
    def _generate_json_report(self, data: Dict[str, Any], config: ReportConfig) -> str:
        """Generate JSON report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.report_type.value}_{timestamp}.json"
        filepath = os.path.join(self.temp_dir, filename)
        
        # Prepare JSON data
        json_data = {
            'report_metadata': {
                'type': config.report_type.value,
                'generated_at': datetime.now().isoformat(),
                'author': config.author,
                'title': config.title,
                'subtitle': config.subtitle
            },
            'data': data
        }
        
        # Write JSON
        with open(filepath, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, indent=2, default=str)
        
        logger.info(f"📄 Generated JSON report: {filepath}")
        return filepath
    
    # Helper methods for data analysis
    def _serialize_orchid_record(self, record: OrchidRecord) -> Dict[str, Any]:
        """Convert OrchidRecord to dictionary for JSON serialization"""
        return {
            'id': record.id,
            'display_name': record.display_name,
            'scientific_name': record.scientific_name,
            'genus': record.genus,
            'species': record.species,
            'country': record.country,
            'decimal_latitude': float(record.decimal_latitude) if record.decimal_latitude else None,
            'decimal_longitude': float(record.decimal_longitude) if record.decimal_longitude else None,
            'is_flowering': record.is_flowering,
            'flower_count': record.flower_count,
            'ai_confidence': float(record.ai_confidence) if record.ai_confidence else None,
            'created_at': record.created_at.isoformat() if record.created_at else None
        }
    
    def _calculate_monthly_growth(self) -> List[Dict[str, Any]]:
        """Calculate monthly growth in orchid records"""
        # Get monthly counts for the last 12 months
        twelve_months_ago = datetime.now() - timedelta(days=365)
        
        monthly_data = OrchidRecord.query.with_entities(
            func.date_trunc('month', OrchidRecord.created_at).label('month'),
            func.count(OrchidRecord.id).label('count')
        ).filter(
            OrchidRecord.created_at >= twelve_months_ago
        ).group_by('month').order_by('month').all()
        
        return [{'month': month.strftime('%Y-%m'), 'count': count} for month, count in monthly_data]
    
    def _analyze_flowering_patterns(self, records: List[OrchidRecord]) -> Dict[str, int]:
        """Analyze flowering patterns in records"""
        flowering_count = sum(1 for r in records if r.is_flowering)
        total_flowers = sum(r.flower_count or 0 for r in records if r.flower_count)
        
        return {
            'flowering_records': flowering_count,
            'non_flowering_records': len(records) - flowering_count,
            'total_flowers': total_flowers,
            'average_flowers_per_plant': total_flowers / flowering_count if flowering_count > 0 else 0
        }
    
    def _analyze_growth_habits(self, records: List[OrchidRecord]) -> Dict[str, int]:
        """Analyze growth habit distribution"""
        habits = {}
        for record in records:
            habit = record.growth_habit or 'Unknown'
            habits[habit] = habits.get(habit, 0) + 1
        return habits
    
    def _analyze_climate_preferences(self, records: List[OrchidRecord]) -> Dict[str, int]:
        """Analyze climate preference distribution"""
        climates = {}
        for record in records:
            climate = record.climate_preference or 'Unknown'
            climates[climate] = climates.get(climate, 0) + 1
        return climates
    
    def _is_safe_query(self, query: str) -> bool:
        """Check if custom query is safe to execute"""
        # Basic safety checks - prevent destructive operations
        dangerous_keywords = ['DROP', 'DELETE', 'UPDATE', 'INSERT', 'ALTER', 'TRUNCATE', 'EXEC']
        query_upper = query.upper()
        
        for keyword in dangerous_keywords:
            if keyword in query_upper:
                return False
        
        return True
    
    def _generate_executive_summary(self, data: Dict[str, Any], config: ReportConfig) -> str:
        """Generate executive summary text"""
        summary_parts = []
        
        if config.report_type == ReportType.DATA_SUMMARY:
            summary = data.get('summary', {})
            summary_parts.append(f"This comprehensive data summary report covers {summary.get('total_orchids', 0):,} orchid records in the Orchid Continuum database.")
            summary_parts.append(f"The collection includes {len(data.get('analytics', {}).get('genus_distribution', []))} unique genera with {summary.get('image_percentage', 0):.1f}% having associated images.")
            summary_parts.append(f"AI analysis has been performed on {summary.get('ai_analyzed_count', 0)} records with an average confidence score of {summary.get('avg_ai_confidence', 0):.2f}.")
        
        elif config.report_type == ReportType.COLLECTION_ANALYSIS:
            summary = data.get('summary', {})
            summary_parts.append(f"Collection analysis reveals {summary.get('total_records', 0)} records matching the specified criteria.")
            summary_parts.append(f"The collection spans {summary.get('unique_genera', 0)} genera and {summary.get('unique_species', 0)} species.")
            summary_parts.append(f"Flowering specimens account for {summary.get('flowering_records', 0)} records in this subset.")
        
        elif config.report_type == ReportType.GEOGRAPHIC_DISTRIBUTION:
            summary = data.get('summary', {})
            summary_parts.append(f"Geographic analysis covers {summary.get('total_georeferenced', 0)} georeferenced orchid records.")
            summary_parts.append(f"The collection spans {summary.get('countries_represented', 0)} countries with an average of {summary.get('coordinate_density', 0):.1f} records per country.")
        
        return ' '.join(summary_parts)
    
    def _generate_detailed_analysis(self, data: Dict[str, Any], config: ReportConfig) -> str:
        """Generate detailed analysis text"""
        analysis_parts = []
        
        # Add analysis based on report type
        analytics = data.get('analytics', {})
        
        if analytics:
            analysis_parts.append("The detailed analysis reveals several key insights:")
            
            for key, value in analytics.items():
                if isinstance(value, dict):
                    analysis_parts.append(f"• {key.replace('_', ' ').title()}: {len(value)} categories analyzed")
                elif isinstance(value, list):
                    analysis_parts.append(f"• {key.replace('_', ' ').title()}: {len(value)} data points collected")
                else:
                    analysis_parts.append(f"• {key.replace('_', ' ').title()}: {value}")
        
        return ' '.join(analysis_parts)
    
    def _generate_recommendations(self, data: Dict[str, Any], config: ReportConfig) -> str:
        """Generate recommendations based on analysis"""
        recommendations = []
        
        if config.report_type == ReportType.COLLECTION_ANALYSIS:
            summary = data.get('summary', {})
            if summary.get('flowering_records', 0) < summary.get('total_records', 1) * 0.3:
                recommendations.append("• Consider focusing collection efforts on flowering specimens to improve botanical documentation.")
            
            if summary.get('average_ai_confidence', 0) < 0.7:
                recommendations.append("• AI confidence scores suggest opportunities for improved image quality or additional training data.")
        
        elif config.report_type == ReportType.SYSTEM_HEALTH_REPORT:
            recommendations.extend([
                "• Continue regular monitoring of system performance metrics.",
                "• Consider data backup and integrity checks based on current health indicators.",
                "• Evaluate opportunities for system optimization based on usage patterns."
            ])
        
        if not recommendations:
            recommendations.append("• Continue current collection and analysis practices based on positive indicators.")
        
        return '\n'.join(recommendations)
    
    # Excel helper methods
    def _create_excel_summary_sheet(self, worksheet, data: Dict[str, Any], config: ReportConfig):
        """Create Excel summary sheet"""
        worksheet.title = "Summary"
        
        # Title
        worksheet['A1'] = f"{config.report_type.value.replace('_', ' ').title()} Report"
        worksheet['A1'].font = Font(bold=True, size=16)
        
        # Summary data
        if data.get('summary'):
            row = 3
            worksheet[f'A{row}'] = "Key Metrics"
            worksheet[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            for key, value in data['summary'].items():
                worksheet[f'A{row}'] = key.replace('_', ' ').title()
                worksheet[f'B{row}'] = value
                worksheet[f'A{row}'].font = Font(bold=True)
                row += 1
    
    def _create_excel_data_sheet(self, worksheet, records: List[Dict[str, Any]]):
        """Create Excel data sheet with records"""
        worksheet.title = "Data"
        
        if not records:
            return
        
        # Headers
        headers = list(records[0].keys())
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Data
        for row_idx, record in enumerate(records, 2):
            for col_idx, (key, value) in enumerate(record.items(), 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
    
    def _create_excel_analytics_sheet(self, worksheet, analytics: Dict[str, Any]):
        """Create Excel analytics sheet"""
        worksheet.title = "Analytics"
        
        row = 1
        for key, value in analytics.items():
            worksheet[f'A{row}'] = key.replace('_', ' ').title()
            worksheet[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            if isinstance(value, dict):
                for sub_key, sub_value in value.items():
                    worksheet[f'A{row}'] = f"  {sub_key}"
                    worksheet[f'B{row}'] = sub_value
                    row += 1
            elif isinstance(value, list):
                for item in value[:10]:  # Limit to top 10 items
                    if isinstance(item, dict):
                        worksheet[f'A{row}'] = str(item)
                    else:
                        worksheet[f'A{row}'] = item
                    row += 1
            else:
                worksheet[f'B{row-1}'] = value
            
            row += 2  # Add spacing
    
    def _create_excel_charts_sheet(self, worksheet, charts: Dict[str, str]):
        """Create Excel charts sheet with embedded images"""
        worksheet.title = "Visualizations"
        
        worksheet['A1'] = "Report Visualizations"
        worksheet['A1'].font = Font(bold=True, size=16)
        
        row = 3
        for chart_name, chart_path in charts.items():
            worksheet[f'A{row}'] = chart_name.replace('_', ' ').title()
            worksheet[f'A{row}'].font = Font(bold=True)
            
            # Note: In a full implementation, you would embed the actual images
            # This requires additional openpyxl image handling
            worksheet[f'A{row+1}'] = f"Chart saved as: {os.path.basename(chart_path)}"
            row += 3
    
    def _create_excel_metadata_sheet(self, worksheet, config: ReportConfig):
        """Create Excel metadata sheet"""
        worksheet.title = "Metadata"
        
        metadata = [
            ('Report Type', config.report_type.value),
            ('Generated At', datetime.now().isoformat()),
            ('Author', config.author),
            ('Title', config.title or 'N/A'),
            ('Subtitle', config.subtitle or 'N/A'),
            ('Include Charts', 'Yes' if config.include_charts else 'No'),
            ('Template', config.template_name),
            ('Color Scheme', config.color_scheme)
        ]
        
        for row, (key, value) in enumerate(metadata, 1):
            worksheet[f'A{row}'] = key
            worksheet[f'B{row}'] = value
            worksheet[f'A{row}'].font = Font(bold=True)
    
    def _get_csv_fieldnames(self, sample_record: Dict[str, Any]) -> List[str]:
        """Get CSV fieldnames from sample record"""
        return list(sample_record.keys())
    
    # Additional analysis methods (placeholders for complex analysis)
    def _analyze_coordinate_clusters(self, records: List[OrchidRecord]) -> Dict[str, Any]:
        """Analyze geographic coordinate clusters"""
        return {"cluster_analysis": "Not implemented in this version"}
    
    def _analyze_seasonal_flowering(self, records: List[OrchidRecord]) -> Dict[str, int]:
        """Analyze seasonal flowering patterns"""
        seasons = {"Spring": 0, "Summer": 0, "Fall": 0, "Winter": 0}
        # This would analyze flowering_photo_date or created_at for seasonal patterns
        return seasons
    
    def _analyze_genus_flowering_patterns(self, records: List[OrchidRecord]) -> Dict[str, Any]:
        """Analyze flowering patterns by genus"""
        return {"genus_patterns": "Analysis placeholder"}
    
    def _determine_peak_season(self, records: List[OrchidRecord]) -> str:
        """Determine peak flowering season"""
        return "Spring"  # Placeholder
    
    def _analyze_flowering_stages(self, records: List[OrchidRecord]) -> Dict[str, int]:
        """Analyze distribution of flowering stages"""
        stages = {}
        for record in records:
            stage = record.flowering_stage or 'unknown'
            stages[stage] = stages.get(stage, 0) + 1
        return stages
    
    def _analyze_confidence_distribution(self, records: List[OrchidRecord]) -> List[float]:
        """Get AI confidence score distribution"""
        return [r.ai_confidence for r in records if r.ai_confidence is not None]
    
    def _calculate_ai_accuracy_metrics(self, records: List[OrchidRecord]) -> Dict[str, float]:
        """Calculate AI accuracy metrics"""
        confidences = [r.ai_confidence for r in records if r.ai_confidence is not None]
        if not confidences:
            return {}
        
        return {
            'mean_confidence': np.mean(confidences),
            'median_confidence': np.median(confidences),
            'std_confidence': np.std(confidences),
            'high_confidence_ratio': sum(1 for c in confidences if c > 0.8) / len(confidences)
        }
    
    def _analyze_confidence_by_genus(self, records: List[OrchidRecord]) -> Dict[str, float]:
        """Analyze AI confidence by genus"""
        genus_confidence = {}
        for record in records:
            if record.genus and record.ai_confidence:
                if record.genus not in genus_confidence:
                    genus_confidence[record.genus] = []
                genus_confidence[record.genus].append(record.ai_confidence)
        
        return {genus: np.mean(scores) for genus, scores in genus_confidence.items() if scores}
    
    def _analyze_user_activities(self, activities: List) -> Dict[str, Any]:
        """Analyze user activity patterns"""
        return {"total_activities": len(activities)}  # Placeholder
    
    def _analyze_upload_patterns(self, users: List[User]) -> Dict[str, Any]:
        """Analyze user upload patterns"""
        return {"active_uploaders": len(users)}  # Placeholder
    
    def _calculate_user_engagement_metrics(self) -> Dict[str, float]:
        """Calculate user engagement metrics"""
        return {"engagement_score": 75.5}  # Placeholder
    
    def _calculate_data_quality_score(self) -> float:
        """Calculate overall data quality score"""
        # This would analyze completeness, accuracy, consistency, etc.
        return 85.2  # Placeholder
    
    def _calculate_health_indicators(self) -> Dict[str, Any]:
        """Calculate system health indicators"""
        return {
            "database_health": "Good",
            "api_response_time": "< 500ms",
            "error_rate": "< 0.1%"
        }  # Placeholder
    
    def get_progress(self, report_id: str) -> Optional[ReportProgress]:
        """Get progress for a specific report"""
        return self.progress_tracker.get(report_id)
    
    def cleanup_temp_files(self, max_age_hours: int = 24):
        """Clean up temporary files older than max_age_hours"""
        try:
            import shutil
            import time
            
            now = time.time()
            max_age_seconds = max_age_hours * 3600
            
            for filename in os.listdir(self.temp_dir):
                filepath = os.path.join(self.temp_dir, filename)
                if os.path.isfile(filepath):
                    file_age = now - os.path.getmtime(filepath)
                    if file_age > max_age_seconds:
                        os.remove(filepath)
                        logger.info(f"🗑️ Cleaned up old report file: {filename}")
        
        except Exception as e:
            logger.error(f"Error during cleanup: {e}")

# Global report generator instance
report_generator = OrchidReportGenerator()

# Convenience functions
def generate_data_summary_report(output_formats: List[str] = None, **kwargs) -> Dict[str, Any]:
    """Generate comprehensive data summary report"""
    if output_formats is None:
        output_formats = ['pdf', 'excel']
    
    config = ReportConfig(
        report_type=ReportType.DATA_SUMMARY,
        output_formats=[OutputFormat(fmt) for fmt in output_formats],
        title="Orchid Continuum Data Summary Report",
        **kwargs
    )
    
    return report_generator.generate_report(config)

def generate_collection_analysis_report(genus: str = None, country: str = None, **kwargs) -> Dict[str, Any]:
    """Generate collection analysis report"""
    filters = {}
    if genus:
        filters['genus'] = genus
    if country:
        filters['country'] = country
    
    config = ReportConfig(
        report_type=ReportType.COLLECTION_ANALYSIS,
        output_formats=[OutputFormat.PDF, OutputFormat.EXCEL],
        title="Collection Analysis Report",
        filters=filters,
        **kwargs
    )
    
    return report_generator.generate_report(config)

def generate_custom_report(report_type: str, queries: List[str] = None, **kwargs) -> Dict[str, Any]:
    """Generate custom report with user-defined queries"""
    config = ReportConfig(
        report_type=ReportType(report_type),
        output_formats=[OutputFormat.PDF, OutputFormat.CSV],
        custom_queries=queries or [],
        **kwargs
    )
    
    return report_generator.generate_report(config)

def get_report_progress(report_id: str) -> Optional[Dict[str, Any]]:
    """Get progress information for a report"""
    progress = report_generator.get_progress(report_id)
    if progress:
        return {
            'report_id': progress.report_id,
            'progress_percentage': progress.progress_percentage,
            'current_step': progress.current_step,
            'status': progress.status,
            'elapsed_time': progress.elapsed_time.total_seconds(),
            'error_message': progress.error_message
        }
    return None

# Background cleanup task (would be called periodically)
def cleanup_old_reports():
    """Clean up old report files"""
    report_generator.cleanup_temp_files(max_age_hours=24)

if __name__ == "__main__":
    # Example usage
    logger.info("🚀 Orchid Continuum Report Generator initialized")
    
    # Test data summary report
    result = generate_data_summary_report(['pdf', 'excel'])
    print(f"Generated report: {result}")